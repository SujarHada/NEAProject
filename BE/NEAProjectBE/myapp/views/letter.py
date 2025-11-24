from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from datetime import datetime
from django.db import transaction
import csv
from drf_spectacular.utils import extend_schema, OpenApiResponse, OpenApiExample, OpenApiParameter
from drf_spectacular.types import OpenApiTypes
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

from ..models import Letter, LetterStatus
from ..serializers import LetterSerializer
from ..permissions import IsViewerOrAdminWithCreateForLetters

class LetterViewSet(viewsets.ModelViewSet):
    queryset = Letter.objects.all().order_by("-created_at")
    serializer_class = LetterSerializer
    permission_classes = [IsViewerOrAdminWithCreateForLetters]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["status"]

    def get_queryset(self):
        queryset = super().get_queryset()
        # Prefetch related items for better performance
        return queryset.prefetch_related('items')
    
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Create a new letter with items"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        letter = serializer.save()
        
        return Response({
            "status": "success",
            "message": "Letter created successfully",
            "data": serializer.data
        }, status=status.HTTP_201_CREATED)

    def list(self, request, *args, **kwargs):
        """Get list of letters with proper response"""
        queryset = self.filter_queryset(self.get_queryset())
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({
                "status": "success",
                "message": "Letters retrieved successfully",
                "data": serializer.data
            })

        serializer = self.get_serializer(queryset, many=True)
        return Response({
            "status": "success",
            "message": "Letters retrieved successfully",
            "data": serializer.data
        })

    def retrieve(self, request, *args, **kwargs):
        """Get single letter details"""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response({
            "status": "success",
            "message": "Letter retrieved successfully",
            "data": serializer.data
        })

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        """Update a letter with items"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        letter = serializer.save()

        return Response({
            "status": "success",
            "message": "Letter updated successfully",
            "data": serializer.data
        })

    def partial_update(self, request, *args, **kwargs):
        """Partially update a letter"""
        return self.update(request, *args, partial=True, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Move letter to bin (soft delete)"""
        instance = self.get_object()
        letter_subject = instance.subject or f"Letter {instance.id}"
        instance.status = LetterStatus.BIN
        instance.save(update_fields=["status", "updated_at"])
        
        return Response({
            "status": "success",
            "message": f"Letter '{letter_subject}' has been moved to bin",
            "id": instance.id
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export letters to CSV, optionally filtered by Nepali date range using 'from' and 'to' (YYYY-MM-DD)"""
        queryset = self.filter_queryset(self.get_queryset())

        def nepali_to_english_digits(s):
            mapping = {
                '०': '0', '१': '1', '२': '2', '३': '3', '४': '4',
                '५': '5', '६': '6', '७': '7', '८': '8', '९': '9'
            }
            return ''.join(mapping.get(ch, ch) for ch in (s or ''))

        start_date = request.query_params.get('from')
        end_date = request.query_params.get('to')

        records = list(queryset)
        if start_date or end_date:
            if not start_date or not end_date:
                return Response({
                    "status": "error",
                    "message": "Both 'from' and 'to' query parameters are required in YYYY-MM-DD format"
                }, status=status.HTTP_400_BAD_REQUEST)

            try:
                datetime.strptime(start_date, '%Y-%m-%d')
                datetime.strptime(end_date, '%Y-%m-%d')
            except Exception:
                return Response({
                    "status": "error",
                    "message": "Invalid date format. Use YYYY-MM-DD for 'from' and 'to'"
                }, status=status.HTTP_400_BAD_REQUEST)

            start_norm = start_date
            end_norm = end_date
            filtered = []
            for letter in records:
                d_norm = nepali_to_english_digits(letter.date)
                if not d_norm:
                    continue
                if len(d_norm) != 10 or d_norm.count('-') != 2:
                    continue
                if start_norm <= d_norm <= end_norm:
                    filtered.append(letter)
            records = filtered

        if not records:
            return Response({
                "status": "error",
                "message": "No letters found to export"
            }, status=status.HTTP_404_NOT_FOUND)

        # Change this line to use utf-8-sig encoding
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response.write('\ufeff')
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
        response['Content-Disposition'] = f'attachment; filename="letters_export_{timestamp}.csv"'
        
        writer = csv.writer(response)
        
        # Write headers
        writer.writerow([
            'सि.नं.', 'च.नं.', 'भौचर क्र. सं.', 'मिति', 'गेटपास नं.', 'रेट पठाउन बाकी',
            'कार्यालय', 'उप कार्यालय',
            'सामानको नाम', 'कम्पनी', 'सिरियल नं.', 'इकाई', 'बुझेको परिमाण पुरानो', 'बुझेको परिमाण नया',
            'बुझ्नेको पुरा नाम', 'थर', 'पद', 'Mobile', 'गाडी नम्बर', 'तयार गर्ने', 'कैफियत'
        ])
        
        def normalize_date(d):
            dn = nepali_to_english_digits(d)
            if dn and dn.count('-') == 2 and len(dn) == 10:
                return dn.replace('-', '.')
            return dn or ''

        for index, letter in enumerate(records, start=1):
            items = list(letter.items.all())
            if not items:
                last_name = (letter.receiver_name or '').strip().split(' ')[-1] if letter.receiver_name else ''
                writer.writerow([
                    index,
                    letter.chalani_no or '',
                    letter.voucher_no or '',
                    normalize_date(letter.date),
                    letter.gatepass_no or '',
                    '-',
                    letter.office_name or '',
                    letter.sub_office_name or '',
                    '',
                    '',
                    '',
                    '',
                    '-',
                    '',
                    letter.receiver_name or '',
                    last_name,
                    letter.receiver_post or '',
                    letter.receiver_phone_number or '',
                    letter.receiver_vehicle_number or '',
                    'Central Store',
                    ''
                ])
            else:
                for it in items:
                    last_name = (letter.receiver_name or '').strip().split(' ')[-1] if letter.receiver_name else ''
                    writer.writerow([
                        index,
                        letter.chalani_no or '',
                        letter.voucher_no or '',
                        normalize_date(letter.date),
                        letter.gatepass_no or '',
                        '-',
                        letter.office_name or '',
                        letter.sub_office_name or '',
                        it.name,
                        it.company,
                        it.serial_number,
                        it.unit_of_measurement,
                        '-',
                        it.quantity,
                        letter.receiver_name or '',
                        last_name,
                        letter.receiver_post or '',
                        letter.receiver_phone_number or '',
                        letter.receiver_vehicle_number or '',
                        'Central Store',
                        ''
                    ])
        
        return response

    @extend_schema(
        request={
            'application/json': {
                'type': 'object',
                'properties': {
                    'start_date': {'type': 'string', 'example': '2082-07-01'},
                    'end_date': {'type': 'string', 'example': '2082-08-03'}
                },
                'required': ['start_date', 'end_date']
            }
        },
        responses={
            200: OpenApiResponse(response=OpenApiTypes.BINARY, description='CSV file'),
            400: OpenApiResponse(response=OpenApiTypes.OBJECT, description='Validation error'),
            404: OpenApiResponse(response=OpenApiTypes.OBJECT, description='No letters found')
        },
        description='Export letters to CSV by Nepali date range using POST body',
        summary='Export Letters CSV by Date Range'
    )
    
    @action(detail=True, methods=['post'])
    def send(self, request, pk=None):
        """Mark letter as sent"""
        instance = self.get_object()
        
        if instance.status == LetterStatus.SENT:
            letter_subject = instance.subject or f"Letter {instance.id}"
            return Response({
                "status": "error",
                "message": f"Letter '{letter_subject}' is already sent"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        instance.status = LetterStatus.SENT
        instance.save(update_fields=["status", "updated_at"])
        
        letter_subject = instance.subject or f"Letter {instance.id}"
        return Response({
            "status": "success",
            "message": f"Letter '{letter_subject}' has been marked as sent",
            "data": LetterSerializer(instance).data
        })

    @action(detail=True, methods=['post'])
    def draft(self, request, pk=None):
        """Mark letter as draft"""
        instance = self.get_object()
        
        if instance.status == LetterStatus.DRAFT:
            letter_subject = instance.subject or f"Letter {instance.id}"
            return Response({
                "status": "error",
                "message": f"Letter '{letter_subject}' is already in draft"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        instance.status = LetterStatus.DRAFT
        instance.save(update_fields=["status", "updated_at"])
        
        letter_subject = instance.subject or f"Letter {instance.id}"
        return Response({
            "status": "success",
            "message": f"Letter '{letter_subject}' has been marked as draft",
            "data": LetterSerializer(instance).data
        })

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        """Restore letter from bin"""
        instance = self.get_object()
        
        if instance.status != LetterStatus.BIN:
            letter_subject = instance.subject or f"Letter {instance.id}"
            return Response({
                "status": "error",
                "message": f"Letter '{letter_subject}' is not in bin"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        instance.status = LetterStatus.DRAFT
        instance.save(update_fields=["status", "updated_at"])
        
        letter_subject = instance.subject or f"Letter {instance.id}"
        return Response({
            "status": "success",
            "message": f"Letter '{letter_subject}' has been restored from bin",
            "data": LetterSerializer(instance).data
        })

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get letter statistics"""
        total_letters = Letter.objects.count()
        draft_letters = Letter.objects.filter(status=LetterStatus.DRAFT).count()
        sent_letters = Letter.objects.filter(status=LetterStatus.SENT).count()
        bin_letters = Letter.objects.filter(status=LetterStatus.BIN).count()
        
        return Response({
            "status": "success",
            "message": "Letter statistics retrieved successfully",
            "data": {
                "total_letters": total_letters,
                "draft_letters": draft_letters,
                "sent_letters": sent_letters,
                "bin_letters": bin_letters
            }
        })
    @extend_schema(
        parameters=[
            OpenApiParameter(
                name='start_date',
                type=OpenApiTypes.STR,
                location=OpenApiParameter.QUERY,
                description='Start date in Nepali format (YYYY-MM-DD)',
                required=True,
                examples=[
                    OpenApiExample(
                        'Example 1',
                        value='2082-07-01'
                    )
                ]
            ),
            OpenApiParameter(
                name='end_date',
                type=OpenApiTypes.STR,
                location=OpenApiParameter.QUERY,
                description='End date in Nepali format (YYYY-MM-DD)',
                required=True,
                examples=[
                    OpenApiExample(
                        'Example 1',
                        value='2082-08-03'
                    )
                ]
            ),
        ],
        description='Get letters by Nepali date range using query parameters',
        summary='Get Letters by Date Range'
    )
    @action(detail=False, methods=['get'], url_path='by-date-range')
    def get_by_date_range(self, request):
        """Get letters by Nepali date range using query parameters"""
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if not start_date or not end_date:
            return Response({
                "status": "error",
                "message": "Both 'start_date' and 'end_date' query parameters are required in YYYY-MM-DD format"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        def nepali_to_english_digits(s):
            mapping = {
                '०': '0', '१': '1', '२': '2', '३': '3', '४': '4',
                '५': '5', '६': '6', '७': '7', '८': '8', '९': '9'
            }
            return ''.join(mapping.get(ch, ch) for ch in (s or ''))
        
        start_norm = nepali_to_english_digits(start_date)
        end_norm = nepali_to_english_digits(end_date)
        
        try:
            # Validate date format
            datetime.strptime(start_norm, '%Y-%m-%d')
            datetime.strptime(end_norm, '%Y-%m-%d')
        except ValueError:
            return Response({
                "status": "error",
                "message": "Invalid date format. Use YYYY-MM-DD for 'start_date' and 'end_date'"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get base queryset
        queryset = self.filter_queryset(self.get_queryset())
        
        # Filter by date range
        filtered_letters = []
        for letter in queryset:
            letter_date_norm = nepali_to_english_digits(letter.date)
            if letter_date_norm and len(letter_date_norm) == 10 and letter_date_norm.count('-') == 2:
                if start_norm <= letter_date_norm <= end_norm:
                    filtered_letters.append(letter.id)
        
        # Apply the filter
        queryset = queryset.filter(id__in=filtered_letters)
        
        # Paginate and return response
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({
                "status": "success",
                "message": f"Letters filtered by date range {start_date} to {end_date}",
                "data": serializer.data
            })

        serializer = self.get_serializer(queryset, many=True)
        return Response({
            "status": "success",
            "message": f"Letters filtered by date range {start_date} to {end_date}",
            "data": serializer.data
        })
    @action(detail=False, methods=['get'], url_path='letter-creation-data')
    def letter_creation_data(self, request):
        def nepali_to_english_digits(s):
            m = {'०':'0','१':'1','२':'2','३':'3','४':'4','५':'5','६':'6','७':'7','८':'8','९':'9'}
            return ''.join(m.get(ch, ch) for ch in (s or ''))
        def english_to_nepali_digits(s):
            m = {'0':'०','1':'१','2':'२','3':'३','4':'४','5':'५','6':'६','7':'७','8':'८','9':'९'}
            return ''.join(m.get(ch, ch) for ch in (s or ''))
        def to_int(s):
            t = ''.join(ch for ch in (s or '') if ch.isdigit())
            return int(t) if t else 0
        last = Letter.objects.order_by('-created_at').first()
        next_chalani = 1
        next_voucher = 1
        if last:
            next_chalani = to_int(nepali_to_english_digits(last.chalani_no)) + 1 if last.chalani_no else 1
            next_voucher = to_int(nepali_to_english_digits(last.voucher_no)) + 1 if last.voucher_no else 1
        today = datetime.now()
        ad_year = today.year
        ad_month = today.month
        bs_year = ad_year + 57
        if ad_month < 7:
            bs_year -= 1
        fy_label = f"{bs_year}/{str(bs_year+1)[-2:]}"
        return Response({
            "chalani_no": str(next_chalani),
            "chalani_no_nepali": english_to_nepali_digits(str(next_chalani)),
            "voucher_no": str(next_voucher),
            "voucher_no_nepali": english_to_nepali_digits(str(next_voucher)),
            "fiscal_year": fy_label
        })

    @action(detail=False, methods=['get'], url_path='export_xlsx')
    def export_xlsx(self, request):
        def nepali_digits(s):
            m = {'0':'०','1':'१','2':'२','3':'३','4':'४','5':'५','6':'६','7':'७','8':'८','9':'९'}
            return ''.join(m.get(ch, ch) for ch in (s or ''))

        def normalize_date(d):
            dn = ''.join({'०':'0','१':'1','२':'2','३':'3','४':'4','५':'5','६':'6','७':'7','८':'8','९':'9'}.get(ch, ch) for ch in (d or ''))
            if dn and dn.count('-') == 2 and len(dn) == 10:
                return dn.replace('-', '.')
            return d or ''

        queryset = self.filter_queryset(self.get_queryset())
        records = list(queryset.prefetch_related('items'))
        if not records:
            return Response({"status": "error", "message": "No letters found to export"}, status=status.HTTP_404_NOT_FOUND)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Letters'

        header_font = Font(name='Noto Sans Devanagari', size=11, bold=True)
        cell_font = Font(name='Noto Sans Devanagari', size=11)

        headers = [
            'सि.नं.', 'च.नं.', 'भौचर क्र. सं.', 'मिति', 'गेटपास नं.', 'रेट पठाउन बाकी',
            'कार्यालय', 'उप कार्यालय',
            'सामानको नाम', 'कम्पनी', 'सिरियल नं.', 'इकाई', 'बुझेको परिमाण पुरानो', 'बुझेको परिमाण नया',
            'बुझ्नेको पुरा नाम', 'थर', 'पद', 'Mobile', 'गाडी नम्बर', 'तयार गर्ने', 'कैफियत'
        ]
        ws.append(headers)
        for i, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            c.font = header_font

        row_idx = 2
        # FIX: Use a single counter for all items across all letters
        serial_counter = 1
        
        for letter in records:
            items = list(letter.items.all())
            if not items:
                last_name = (letter.receiver_name or '').strip().split(' ')[-1] if letter.receiver_name else ''
                row = [
                    serial_counter,  # Use the continuous counter
                    str(letter.chalani_no or ''),
                    str(letter.voucher_no or ''),
                    normalize_date(letter.date),
                    str(letter.gatepass_no or ''),
                    '-',
                    letter.office_name or '',
                    letter.sub_office_name or '',
                    '', '', '', '', '-', '',
                    letter.receiver_name or '',
                    last_name,
                    letter.receiver_post or '',
                    nepali_digits(str(letter.receiver_phone_number or '')),
                    nepali_digits(str(letter.receiver_vehicle_number or '')),
                    'Central Store',
                    ''
                ]
                ws.append(row)
                for col in range(1, len(headers)+1):
                    ws.cell(row=row_idx, column=col).font = cell_font
                    ws.cell(row=row_idx, column=col).number_format = '@'
                row_idx += 1
                serial_counter += 1
            else:
                for it in items:
                    last_name = (letter.receiver_name or '').strip().split(' ')[-1] if letter.receiver_name else ''
                    row = [
                        serial_counter,
                        str(letter.chalani_no or ''),
                        str(letter.voucher_no or ''),
                        normalize_date(letter.date),
                        str(letter.gatepass_no or ''),
                        '-',
                        letter.office_name or '',
                        letter.sub_office_name or '',
                        it.name,
                        it.company,
                        it.serial_number,
                        it.unit_of_measurement,
                        '-',
                        str(it.quantity),
                        letter.receiver_name or '',
                        last_name,
                        letter.receiver_post or '',
                        nepali_digits(str(letter.receiver_phone_number or '')),
                        nepali_digits(str(letter.receiver_vehicle_number or '')),
                        'Central Store',
                        ''
                    ]
                    ws.append(row)
                    for col in range(1, len(headers)+1):
                        ws.cell(row=row_idx, column=col).font = cell_font
                        ws.cell(row=row_idx, column=col).number_format = '@'
                    row_idx += 1
                    serial_counter += 1

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        ts = datetime.now().strftime('%Y-%m-%d_%H-%M')
        resp = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="letters_export_{ts}.xlsx"'
        return resp

    @extend_schema(
        request={
            'application/json': {
                'type': 'object',
                'properties': {
                    'start_date': {'type': 'string', 'example': '2082-07-01'},
                    'end_date': {'type': 'string', 'example': '2082-08-03'}
                },
                'required': ['start_date', 'end_date']
            }
        },
        responses={
            200: OpenApiResponse(response=OpenApiTypes.BINARY, description='XLSX file'),
            400: OpenApiResponse(response=OpenApiTypes.OBJECT, description='Validation error'),
            404: OpenApiResponse(response=OpenApiTypes.OBJECT, description='No letters found')
        },
        description='Export letters to XLSX by Nepali date range using POST body',
        summary='Export Letters XLSX by Date Range'
    )
    @action(detail=False, methods=['post'], url_path='export_xlsx_by_date')
    def export_xlsx_by_date(self, request):
        def en_digits(s):
            m = {'०':'0','१':'1','२':'2','३':'3','४':'4','५':'5','६':'6','७':'7','८':'8','९':'9'}
            return ''.join(m.get(ch, ch) for ch in (s or ''))
        def np_digits(s):
            m = {'0':'०','1':'१','2':'२','3':'३','4':'४','5':'५','6':'६','7':'७','8':'८','9':'९'}
            return ''.join(m.get(ch, ch) for ch in (s or ''))
        def normalize_date(d):
            dn = en_digits(d)
            if dn and dn.count('-') == 2 and len(dn) == 10:
                return dn.replace('-', '.')
            return d or ''

        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        if not start_date or not end_date:
            return Response({
                "status": "error",
                "message": "Both 'start_date' and 'end_date' fields are required in YYYY-MM-DD format"
            }, status=status.HTTP_400_BAD_REQUEST)
        start_norm = en_digits(start_date)
        end_norm = en_digits(end_date)
        try:
            datetime.strptime(start_norm, '%Y-%m-%d')
            datetime.strptime(end_norm, '%Y-%m-%d')
        except Exception:
            return Response({
                "status": "error",
                "message": "Invalid date format. Use YYYY-MM-DD for 'start_date' and 'end_date'"
            }, status=status.HTTP_400_BAD_REQUEST)

        queryset = self.filter_queryset(self.get_queryset())
        records = []
        for letter in queryset.prefetch_related('items'):
            d_norm = en_digits(letter.date)
            if not d_norm:
                continue
            if len(d_norm) != 10 or d_norm.count('-') != 2:
                continue
            if start_norm <= d_norm <= end_norm:
                records.append(letter)

        if not records:
            return Response({
                "status": "error",
                "message": "No letters found to export"
            }, status=status.HTTP_404_NOT_FOUND)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Letters'
        header_font = Font(name='Noto Sans Devanagari', size=11, bold=True)
        cell_font = Font(name='Noto Sans Devanagari', size=11)
        headers = [
            'सि.नं.', 'च.नं.', 'भौचर क्र. सं.', 'मिति', 'गेटपास नं.', 'रेट पठाउन बाकी',
            'कार्यालय', 'उप कार्यालय',
            'सामानको नाम', 'कम्पनी', 'सिरियल नं.', 'इकाई', 'बुझेको परिमाण पुरानो', 'बुझेको परिमाण नया',
            'बुझ्नेको पुरा नाम', 'थर', 'पद', 'Mobile', 'गाडी नम्बर', 'तयार गर्ने', 'कैफियत'
        ]
        ws.append(headers)
        for i, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            c.font = header_font
        row_idx = 2

        # FIX: Use a single continuous counter for all items across all letters
        serial_counter = 1

        for letter in records:
            items = list(letter.items.all())
            if not items:
                last_name = (letter.receiver_name or '').strip().split(' ')[-1] if letter.receiver_name else ''
                row = [
                    serial_counter,  # Use continuous counter instead of index
                    str(letter.chalani_no or ''),
                    str(letter.voucher_no or ''),
                    normalize_date(letter.date),
                    str(letter.gatepass_no or ''),
                    '-',
                    letter.office_name or '',
                    letter.sub_office_name or '',
                    '', '', '', '', '-', '',
                    letter.receiver_name or '',
                    last_name,
                    letter.receiver_post or '',
                    np_digits(str(letter.receiver_phone_number or '')),
                    np_digits(str(letter.receiver_vehicle_number or '')),
                    'Central Store',
                    ''
                ]
                ws.append(row)
                for col in range(1, len(headers)+1):
                    ws.cell(row=row_idx, column=col).font = cell_font
                    ws.cell(row=row_idx, column=col).number_format = '@'
                row_idx += 1
                serial_counter += 1  # Increment counter for each row
            else:
                for it in items:
                    last_name = (letter.receiver_name or '').strip().split(' ')[-1] if letter.receiver_name else ''
                    row = [
                        serial_counter,  # Use continuous counter instead of index
                        str(letter.chalani_no or ''),
                        str(letter.voucher_no or ''),
                        normalize_date(letter.date),
                        str(letter.gatepass_no or ''),
                        '-',
                        letter.office_name or '',
                        letter.sub_office_name or '',
                        it.name,
                        it.company,
                        it.serial_number,
                        it.unit_of_measurement,
                        '-',
                        str(it.quantity),
                        letter.receiver_name or '',
                        last_name,
                        letter.receiver_post or '',
                        np_digits(str(letter.receiver_phone_number or '')),
                        np_digits(str(letter.receiver_vehicle_number or '')),
                        'Central Store',
                        ''
                    ]
                    ws.append(row)
                    for col in range(1, len(headers)+1):
                        ws.cell(row=row_idx, column=col).font = cell_font
                        ws.cell(row=row_idx, column=col).number_format = '@'
                    row_idx += 1
                    serial_counter += 1  # Increment counter for each row

        info = wb.create_sheet('Instructions')
        info_text = (
            'यो फाइल Unicode (UTF-8) मा तयार गरिएको छ।\n'
            f"मिति दायरा: {start_norm} देखि {end_norm} सम्म।\n"
            'गाडी नम्बर र Mobile स्तम्भहरूमा नेपाली अंक (०–९) प्रयोग गरिएको छ।\n'
            'Google Sheets मा Noto Sans Devanagari फन्ट प्रयोग गर्नुहोस्।\n'
            'Excel मा Noto Sans Devanagari वा Mangal फन्ट इन्स्टल गरिएको हुनुपर्छ।\n'
            'Preeti जस्तो legacy फन्ट प्रयोग नगर्नुहोस्; Unicode मात्र राख्नुहोस्।'
        )
        info.cell(row=1, column=1, value=info_text)
        info.column_dimensions['A'].width = 120

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        ts = datetime.now().strftime('%Y-%m-%d_%H-%M')
        resp = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="letters_export_{ts}_range.xlsx"'
        return resp
